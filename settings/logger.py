import os
import logging
from datetime import datetime

from django.conf import settings
from openpyxl import Workbook


class ExcelFileHandler(logging.FileHandler):
    """Handler for logging to excel file."""

    def __init__(self, filename):
        filepath = os.path.join(settings.BASE_DIR, filename)
        if not os.path.isfile(filepath):
            open(filepath, 'a').close()
        super().__init__(filepath)
        self.setLevel(logging.INFO)
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.workbook.create_sheet(title='INFO', index=0)
        self.workbook.create_sheet(title='WARNING', index=1)
        self.workbook.create_sheet(title='ERROR', index=2)
        self.workbook.create_sheet(title='CRITICAL', index=3)

    def emit(self, record):
        try:
            msg = self.format(record)
            sheet_name = record.levelname
            sheet = self.workbook[sheet_name]

            created_datetime = datetime.fromtimestamp(record.created)

            row = [
                record.levelname, 
                created_datetime.strftime("%Y-%m-%d"), 
                created_datetime.strftime("%H:%M:%S"), 
                msg
            ]
            sheet.append(row)
            self.workbook.save(self.baseFilename)
        except (KeyboardInterrupt, SystemExit):
            raise
        except Exception:
            self.handleError(record)
